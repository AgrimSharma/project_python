import pdb
from openpyxl.reader.excel import load_workbook

from service_centre.models import Shipment

cust_struct = {
    11007:{
        0:'order_number', #Suborder Number
        1:'shipper',
        2: 'consignee',
        3: 'consignee_address1',
        4: 'consignee_address2',
        5: 'destination_city',
        6: 'state',
        7: 'pincode',
        8: 'telephone',
        9: 'mobile',
        10: 'ShippingTeleNo2',
        11: 'declared_value',
        12: 'product_type',
        13: 'customer_name',
        14: 'collectable_value',
        15: 'actual_weight',
        16: 'Dummy', #Dummy
        17: 'CourierCode', #?
        18: 'ProductCode', #?
        19: 'item_description',
        20: 'ItemCode', #?
        21: 'airwaybill_number',
        22: 'PickupLocation', #?
        23: 'sub_customer_code',
    }
}

ecomm_struct = {
    'airwaybill_number':0,
    'order_number':1,
    'product_type':2,
    'shipper':3,
    'consignee':4,
    'consignee_address1':5,
    'consignee_address2':6,
    'consignee_address3':7,
    'destination_city':8,
    'pincode':9,
    'state':10,
    'mobile':11,
    'telephone':12,
    'item_description':13,
    'pieces':14,
    'collectable_value':15,
    'declared_value':16,
    'actual_weight':17,
    'volumetric_weight':18,
    'length':19,
    'breadth':20,
    'height':21,
    'sub_customer_code':22
    }

ecomm_col_count = len(ecomm_struct.keys())

def get_index_order(cid):
    """
    This function will return a list of positions.
    [1, 3, 4, 5, 6, 8, 10, 9, 12, 11, None, 16, 2, None, 15, 17, None, None, None, 13, None, 0, None, 22]
    each value at an index says, " my 'index' column is in 'value' position in ecomm structure."
    """
    total_cols = len(cust_struct[cid].values())
    index_order = [None] * ecomm_col_count
    for col in range(total_cols):
        index = ecomm_struct.get(cust_struct[cid][col], None)
        if index is None:
            continue
        else:
            index_order[col] = index

    return index_order

def get_data_in_ecomm_order(upload_file, cid):
    file_contents = open(upload_file, 'rb')
    if not file_contents:
        return None

    wb = load_workbook(file_contents, use_iterators=True)
    sh = wb.get_sheet_by_name(wb.get_sheet_names()[0])

    # get all columns data
    num_cols = sh.get_highest_column()
    num_rows = sh.get_highest_row()

    # data matrix corresponding to ecomm structure
    data_matrix = [[None] * ecomm_col_count] * (num_rows - 1)
    index_order = get_index_order(cid) # list: order in which data to be entered

    sheet_matrix = []
    for row in sh.iter_rows():
        data = [cell.internal_value for cell in row]
        sheet_matrix.append(data)

    def get_column(n):
        return [row[n] for row in sheet_matrix][1:]

    pdb.set_trace()
    print 'col count :', num_cols
    for col in range(num_cols):
        print col
        col_data = get_column(col)
        print col_data
        ecomm_index = index_order[col]
        if ecomm_index:
            data_matrix[ecomm_index] = col_data
    print 'for loop end'
    return data_matrix

def process_customer_pickup(cid):
    # read the uploaded excel file: based on the customer rearrange
    # the data matrix so that it will match with ecomm data structure
    data_matrix = get_data_in_ecomm_order(request.FILES['upload_file'], content, cid)
    if data_matrix is None:
        return False

    # process the data

    # send back the error list

def auto_upload_file():
    print " inside "
    pid=1
    dup_awb = []
    awb_overweight=[]
    subCustomers_list=[]
    if request.POST:
            upload_file = request.FILES['upload_file']
            file_contents = upload_file.read()
            if file_contents:
                import_wb = xlrd.open_workbook(file_contents=file_contents)
                import_sheet = import_wb.sheet_by_index(0)
                for a in range(1, import_sheet.nrows):
                   for field in [3,4,5,8,9,10,11,13,17,19,20,21]:
                       field_data = import_sheet.cell_value(rowx=a, colx=field)
                       val = field_data.encode('utf-8') if isinstance(field_data,unicode)  else field_data
                       if field == 17:
                          if float(val) <= 0.0:
                             return HttpResponse("Airwaybill with incorrect weight found %s-%s"%(a,field))
                       if not val:
                              return HttpResponse("Field left blank - file could not be uploaded %s-%s"%(a,field))
                   airwaybill_num = import_sheet.cell_value(rowx=a, colx=0)
                   coll_val = import_sheet.cell_value(rowx=a, colx=15)
                 #  return HttpResponse("Field left blank - file could not be uploaded %s-%s"%(airwaybill_num,coll_val))
                   if airwaybill_num:
                     if str(airwaybill_num)[0] in ['7','8','9'] and float(coll_val) <= 0.0:
                            return HttpResponse("COD shipment found with 0 collectible value")
                     if Shipment.objects.filter(airwaybill_number=airwaybill_num):
                       awb_num = Shipment.objects.get(airwaybill_number=airwaybill_num)
                       #if (awb_num.status <> 0 and awb_num.status <> 1 and awb_num.return_shipment <> 0):
                       if (awb_num.status >=2 or awb_num.return_shipment > 0):
                            return HttpResponse("Used Air waybill entered, please recheck file before uploading.")
                       #return HttpResponse("-----%s" % (awb_num.return_shipment <> 0))
                   #except:
                   #     pass

                   if airwaybill_num not in dup_awb:
                          dup_awb.append(airwaybill_num)
                   else:
                       return HttpResponse("Recheck file, duplicate airwaybill number found")

                reverse_pickup=0

                for rx in range(1, import_sheet.nrows):
                    sub_customer_id = import_sheet.cell_value(rowx=rx, colx=22)
                    if Shipper.objects.filter(id=sub_customer_id):
                        sub_customer = Shipper.objects.get(id=sub_customer_id)
                        if not sub_customer in subCustomers_list:
                            subCustomers_list.append(sub_customer)
                    else:
                        return HttpResponse("Subcustomer not found! Please check again.")

                for i in enumerate(subCustomers_list):
                    print "Submerchants are",i

                pickup_dict = {}

                for subcust in subCustomers_list:
                     try:
                       pincode = Pincode.objects.get(pincode = int(subcust.address.pincode))
                     except:
                      return HttpResponse("Pincode does not exists for this subcustomer")
                     print "pincode",pincode.service_center
                     pickup = PickupRegistration(customer_code = subcust.customer,subcustomer_code=subcust,pickup_time=now,pickup_date=now,mode_id=1,customer_name=subcust.name,address_line1=subcust.address.address1,address_line2=subcust.address.address2,pincode=subcust.address.pincode,address_line3=subcust.address.address3,address_line4=subcust.address.address4,mobile=0,telephone=0,pieces=4,actual_weight=1.2,volume_weight=2.1,service_centre=pincode.service_center)
                     pickup.save()
                     pickup_dict[subcust.id] = pickup
                     subcus = pickup_dict.keys()
                     #code for matching with scheduled
                     scheduled_pickup = PickupSchedulerRegistration.objects.filter(status = 0, service_centre = request.user.employeemaster.service_centre, subcustomer_code__id__in=subcus)
                     for a in scheduled_pickup:
                         a.status =1
                         a.save()
                for rx in range(1, import_sheet.nrows):
                    airwaybill_num = import_sheet.cell_value(rowx=rx, colx=0)
                    pickup = pickup_dict[int(import_sheet.cell_value(rowx=rx, colx=22))]
                    try:
                       awb_num = AirwaybillNumbers.objects.get(airwaybill_number=airwaybill_num)
                       awb_num.status=1
                       awb_num.save()
                    except:
                        return HttpResponse("Wrong airwaybill Number")
                    order_num = import_sheet.cell_value(rowx=rx, colx=1)
                    product_type = import_sheet.cell_value(rowx=rx, colx=2)
                    product_type = product_type.lower()
                    if str(airwaybill_num)[0] in ["1","2","3"]:
                           product_type="ppd"
                    elif str(airwaybill_num)[0] in ["7","8","9"]:
                           product_type="cod"

                    shipper = import_sheet.cell_value(rowx=rx, colx=3)
                    shipper = pickup.customer_code
                    if awb_num.awbc_info.get().customer.code <> "32012":
                       if shipper <> awb_num.awbc_info.get().customer:
                         return HttpResponse("Airwaybill does not belong to this Customer, please recheck")

                    consignee = import_sheet.cell_value(rowx=rx, colx=4)
                    consignee_address1 = import_sheet.cell_value(rowx=rx, colx=5)
                    consignee_address2 = import_sheet.cell_value(rowx=rx, colx=6)
                    consignee_address3 = import_sheet.cell_value(rowx=rx, colx=7)

                    destination_city = import_sheet.cell_value(rowx=rx, colx=8)
                    pincode = import_sheet.cell_value(rowx=rx, colx=9)
                    state = import_sheet.cell_value(rowx=rx, colx=10)
                    mobile = import_sheet.cell_value(rowx=rx, colx=11)
                    telephone = import_sheet.cell_value(rowx=rx, colx=12)
                    item_description = import_sheet.cell_value(rowx=rx, colx=13)
                    pieces = import_sheet.cell_value(rowx=rx, colx=14)
                    collectable_value=import_sheet.cell_value(rowx=rx, colx=15)
                    declared_value=import_sheet.cell_value(rowx=rx, colx=16)
                    actual_weight = import_sheet.cell_value(rowx=rx, colx=17)
                    volumetric_weight = import_sheet.cell_value(rowx=rx, colx=18)
                    length = import_sheet.cell_value(rowx=rx, colx=19)
                    breadth = import_sheet.cell_value(rowx=rx, colx=20)
                    height = import_sheet.cell_value(rowx=rx, colx=21)
                    order_num = repr(import_sheet.cell_value(rowx=rx, colx=1))
                    if order_num.replace(".", "", 1).isdigit():
                       order_num = int(float(order_num))
                    elif 'e+' in str(order_num):
                       order_num = int(float(order_num))
                    else:
                       order_num = import_sheet.cell_value(rowx=rx, colx=1)

                    if length == "":
                       length = 0.0
                    if breadth == "":
                       breadth = 0.0
                    if height == "":
                       height = 0.0
                    if actual_weight == "":
                       actual_weight = 0.0
                    if not (isinstance(volumetric_weight, float) or isinstance(volumetric_weight, int)):
                      if volumetric_weight.strip() == "":

                         volumetric_weight = 0.0
                    if ((actual_weight > 10.0) or (volumetric_weight > 10.0)):
                            a = str(airwaybill_num)+"("+str(max(actual_weight,volumetric_weight))+"Kgs)"
                            awb_overweight.append(a)
                    if collectable_value == "":
                       collectable_value = 0.0
                    else:
                        try:
                            int(collectable_value)
                        except ValueError:
                            collectable_value = collectable_value.replace(",", "")
                    if declared_value == "":
                       declared_value = 0.0
                    else:
                       try:
                            int(declared_value)
                       except ValueError:
                            declared_value = declared_value.replace(",", "")
                    if mobile == "":
                       mobile = 0

                    if pincode == "":
                       pincode = 0.0
                       tt_duration = 0
                    else:

                        origin_pincode=pickup.pincode
                        if not Pincode.objects.filter(pincode=pincode):
                            pincode = 0.0
                            tt_duration = 0
                        try:
                            pincode1 = Pincode.objects.get(pincode=origin_pincode)
                            origin_service_centre = pickup.service_centre
                            sctmg = ServiceCenterTransitMasterGroup.objects.get(service_center=origin_service_centre)
                            dest_pincode=Pincode.objects.get(pincode=pincode)
                            dest_service_centre = dest_pincode.service_center
                            tt_duration = 0

                            transit_time = TransitMaster.objects.filter(transit_master=sctmg.transit_master_group,
                                                    dest_service_center=dest_service_centre)
                            if transit_time:
                                transit_time = transit_time[0]

                                transit_time_cutoff =  TransitMasterCutOff.objects.filter(transit_master_orignal=sctmg.transit_master_group,
                                                    transit_master_dest=dest_service_centre.servicecentertransitmastergroup_set.get())
                                if not transit_time_cutoff:
                                    cutoff = datetime.datetime.strptime(transit_time.cutoff_time,"%H:%M")
                                else:
                                    cutoff = datetime.datetime.strptime(transit_time_cutoff[0].cutoff_time,"%H:%M")

                                tt_duration=int(transit_time.duration)
                            else:
                                tt_duration=0
                        except ValueError:
                            tt_duration=0
                   #return HttpResponse(tt_duration)
                    servicecentre = None
                    if pincode:
                        try:
                          pincode_sc = Pincode.objects.get(pincode=pincode)
                          servicecentre = pincode_sc.service_center
                        except:
                          pincode = ""
                          servicecentre = None
                    expected_dod = None
                    if tt_duration <> 0:
                         if now.time() > cutoff.time():
                             tt_duration+=1
                         expected_dod = now + datetime.timedelta(days=tt_duration)
                         try:
                             HolidayMaster.objects.get(date=expected_dod.date())
                             expected_dod = expected_dod + datetime.timedelta(days=1)
                         except:
                             pass

                    try:
                     shipment = Shipment.objects.filter(airwaybill_number=airwaybill_num, status__in=[0,1]).update(order_number=str(order_num), current_sc=pickup.service_centre,product_type=product_type, shipper=shipper, pickup=pickup, reverse_pickup=reverse_pickup, consignee=consignee, consignee_address1 = consignee_address1, consignee_address2 = consignee_address2 , consignee_address3 = consignee_address3, destination_city=destination_city, pincode=pincode, state=state, mobile=mobile, telephone=telephone, item_description=item_description, pieces=pieces, collectable_value=collectable_value, declared_value=declared_value, actual_weight=actual_weight, volumetric_weight=volumetric_weight, length=length, breadth=breadth, height=height, service_centre = servicecentre, original_dest = servicecentre, expected_dod = expected_dod)
                     shipment = Shipment.objects.get(airwaybill_number=airwaybill_num)
                    except:
                      shipment = Shipment(airwaybill_number=int(airwaybill_num), current_sc=pickup.service_centre, order_number=str(order_num), product_type=product_type, shipper=shipper, pickup=pickup, reverse_pickup=reverse_pickup, consignee=consignee, consignee_address1 = consignee_address1, consignee_address2 = consignee_address2 , consignee_address3 = consignee_address3, destination_city=destination_city, pincode=pincode, state=state, mobile=mobile, telephone=telephone, item_description=item_description, pieces=pieces, collectable_value=collectable_value, declared_value=declared_value, actual_weight=actual_weight, volumetric_weight=volumetric_weight, length=length, breadth=breadth, height=height, service_centre = servicecentre, original_dest = servicecentre, expected_dod = expected_dod)
                      shipment.save()
                   # if pincode:
                   #     try:
                   #       pincode = Pincode.objects.get(pincode=shipment.pincode)
                   #       servicecentre = pincode.service_center
                   #       shipment.service_centre = servicecentre
                   #       shipment.original_dest = servicecentre
                   #       shipment.save()
                   #     except:
                   #       pincode = ""

                   # if tt_duration <> 0:
                   #      if shipment.added_on.time() > cutoff.time():
                   #          tt_duration+=1
                   #      expected_dod = shipment.added_on + datetime.timedelta(days=tt_duration)
                   #      try:
                   #          HolidayMaster.objects.get(date=expected_dod.date())
                   #          expected_dod = expected_dod + datetime.timedelta(days=1)
                   #      except:
                   #          pass
                   #      shipment.expected_dod=expected_dod
                   #      shipment.save()
                    if shipment:
                         history_update(shipment, 0, request)
                    tmp_count=Shipment.objects.filter(pickup=pickup.id).count()
                    pickup.pieces=tmp_count;
                    pickup.status=0
                    pickup.save()
            else:
              pass

            group = Group.objects.get(name="Customer Service")
            a=0
            if request.user.groups.filter(name="Customer Service").exists():
              pickup = PickupRegistration.objects.filter().order_by('-added_on')
              a=1
            if request.user.employeemaster.user_type in ["Staff", "Supervisor", "Sr Supervisor"]:
               pickup = PickupRegistration.objects.filter(service_centre=request.user.employeemaster.service_centre, status=0)
            else:
               pickup = PickupRegistration.objects.filter(status=0).order_by('-added_on')
            msg = ""
            if awb_overweight:
              msg = "Following Air Waybills have weight more than 10 kgs, PLEASE CONFIRM !!!, and incase of mismatch re-upload the file:\n"+"\n".join(['%s' % ship for ship in awb_overweight])
            customer=Customer.objects.all()
            destination= ServiceCenter.objects.all()
            return render_to_response("pickup/pickupdashboard.html",
                              {'pickup':pickup,
                               'msg':msg,
                               'a':a,
                               'customer':customer,
                               'sc':destination},
                               context_instance=RequestContext(request))
    else:
     return render_to_response('pickup/auto_upload_file.html',
                               {'pid':pid,
                                },context_instance=RequestContext(request))


